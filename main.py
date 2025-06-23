import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import datetime
import planning
from tasks import TaskManager
import openpyxl

class PlanningApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Planning Manager")
        self.geometry("1200x700")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

        self.task_manager = TaskManager()
        self.current_year = datetime.date.today().year
        self.current_month = datetime.date.today().month

        self.persons = planning.PERSONS
        self.person_colors = planning.PERSON_COLORS
        self.shifts = planning.SHIFTS

        self.rest_days = {p: [] for p in self.persons}  # jours de repos

        self.create_widgets()
        self.refresh_planning_table()
        self.refresh_task_list()

    def create_widgets(self):
        # Navigation haut
        nav_frame = ctk.CTkFrame(self)
        nav_frame.pack(fill="x", pady=5)

        self.btn_prev_month = ctk.CTkButton(nav_frame, text="<< Mois précédent", command=self.prev_month)
        self.btn_prev_month.pack(side="left", padx=10)

        self.lbl_month = ctk.CTkLabel(nav_frame, text="", font=ctk.CTkFont(size=16, weight="bold"))
        self.lbl_month.pack(side="left", padx=10)

        self.btn_next_month = ctk.CTkButton(nav_frame, text="Mois suivant >>", command=self.next_month)
        self.btn_next_month.pack(side="left", padx=10)

        # Cadre planning
        self.table_frame = ctk.CTkFrame(self)
        self.table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.canvas = tk.Canvas(self.table_frame, bg="#222222", highlightthickness=0)
        self.v_scroll = tk.Scrollbar(self.table_frame, orient="vertical", command=self.canvas.yview)
        self.h_scroll = tk.Scrollbar(self.table_frame, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set, xscrollcommand=self.h_scroll.set)

        self.v_scroll.pack(side="right", fill="y")
        self.h_scroll.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.inner_frame = ctk.CTkFrame(self.canvas)
        self.canvas.create_window((0, 0), window=self.inner_frame, anchor="nw")
        self.inner_frame.bind("<Configure>", self.on_frame_configure)

        # Bas: gestion tâches + repos
        bottom_frame = ctk.CTkFrame(self)
        bottom_frame.pack(fill="x", padx=10, pady=5)

        # Formulaire tâche + liste tâche
        task_frame = ctk.CTkFrame(bottom_frame)
        task_frame.pack(side="left", padx=10, pady=10, fill="y")

        ctk.CTkLabel(task_frame, text="Ajouter une tâche", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(0,10))

        # Combo tâches déjà connues (extrait noms uniques)
        self.task_names = []
        self.combo_task_name = ctk.CTkComboBox(task_frame, values=self.task_names, width=200)
        self.combo_task_name.pack(pady=5)
        self.combo_task_name.set("Sélectionnez une tâche")

        # Bouton ajouter nouvelle tâche (popup)
        btn_new_task = ctk.CTkButton(task_frame, text="Nouvelle tâche...", command=self.add_new_task_popup)
        btn_new_task.pack(pady=5)

        self.entry_task_duration = ctk.CTkEntry(task_frame, placeholder_text="Durée (min)", width=200)
        self.entry_task_duration.pack(pady=5)

        ctk.CTkLabel(task_frame, text="Personne:").pack(anchor="w", pady=(10,0))
        self.combo_person = ctk.CTkComboBox(task_frame, values=self.persons, width=200)
        self.combo_person.pack(pady=5)
        self.combo_person.set(self.persons[0])

        ctk.CTkLabel(task_frame, text="Date (AAAA-MM-JJ):").pack(anchor="w", pady=(10,0))
        self.entry_date = ctk.CTkEntry(task_frame, width=200)
        self.entry_date.pack(pady=5)
        self.entry_date.insert(0, datetime.date.today().isoformat())

        ctk.CTkLabel(task_frame, text="Shift:").pack(anchor="w", pady=(10,0))
        self.combo_shift = ctk.CTkComboBox(task_frame, values=[f"{s[0]}-{s[1]}" for s in self.shifts], width=200)
        self.combo_shift.pack(pady=5)
        self.combo_shift.set(f"{self.shifts[0][0]}-{self.shifts[0][1]}")

        btn_add_task = ctk.CTkButton(task_frame, text="Ajouter tâche", command=self.add_task)
        btn_add_task.pack(pady=10)

        # Liste des tâches existantes
        ctk.CTkLabel(task_frame, text="Tâches assignées", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(20,5))

        self.tree_tasks = ttk.Treeview(task_frame, columns=("name", "duration", "person", "date", "shift"), show="headings", height=10)
        self.tree_tasks.pack(fill="y")
        self.tree_tasks.heading("name", text="Nom")
        self.tree_tasks.heading("duration", text="Durée (min)")
        self.tree_tasks.heading("person", text="Personne")
        self.tree_tasks.heading("date", text="Date")
        self.tree_tasks.heading("shift", text="Shift")

        # Bouton supprimer tâche sélectionnée
        btn_del_task = ctk.CTkButton(task_frame, text="Supprimer tâche sélectionnée", command=self.delete_selected_task)
        btn_del_task.pack(pady=10)

        # Rest days gestion simple
        rest_frame = ctk.CTkFrame(bottom_frame)
        rest_frame.pack(side="left", padx=20, pady=10)

        ctk.CTkLabel(rest_frame, text="Gestion des jours de repos", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(0,10))

        ctk.CTkLabel(rest_frame, text="Personne:").pack(anchor="w")
        self.combo_rest_person = ctk.CTkComboBox(rest_frame, values=self.persons, width=150)
        self.combo_rest_person.pack(pady=5)
        self.combo_rest_person.set(self.persons[0])

        ctk.CTkLabel(rest_frame, text="Jours de repos (YYYY-MM-DD), séparés par ;").pack(anchor="w", pady=(5,0))
        self.entry_rest_days = ctk.CTkEntry(rest_frame, width=250)
        self.entry_rest_days.pack(pady=5)
        self.entry_rest_days.insert(0, "")

        btn_save_rest = ctk.CTkButton(rest_frame, text="Enregistrer repos", command=self.save_rest_days)
        btn_save_rest.pack(pady=10)

        # Exporter planning
        btn_export = ctk.CTkButton(bottom_frame, text="Exporter planning Excel", command=self.export_excel)
        btn_export.pack(side="right", padx=20, pady=10)

    def add_new_task_popup(self):
        popup = ctk.CTkToplevel(self)
        popup.title("Nouvelle tâche")
        popup.geometry("300x150")

        ctk.CTkLabel(popup, text="Nom tâche:").pack(pady=5)
        entry_name = ctk.CTkEntry(popup)
        entry_name.pack(pady=5)

        ctk.CTkLabel(popup, text="Durée (min):").pack(pady=5)
        entry_duration = ctk.CTkEntry(popup)
        entry_duration.pack(pady=5)

        def save_new_task():
            name = entry_name.get().strip()
            dur_str = entry_duration.get().strip()
            if not name or not dur_str:
                messagebox.showerror("Erreur", "Veuillez remplir tous les champs")
                return
            try:
                dur = int(dur_str)
                if dur <= 0:
                    raise ValueError()
            except:
                messagebox.showerror("Erreur", "Durée invalide")
                return
            # Ajouter tâche à task_manager sans assignation
            self.task_manager.add_task(name, dur, None)
            self.refresh_task_names()
            popup.destroy()

        btn_save = ctk.CTkButton(popup, text="Ajouter", command=save_new_task)
        btn_save.pack(pady=10)

    def refresh_task_names(self):
        # Extraire noms uniques des tâches sans doublons
        names = list({t["name"] for t in self.task_manager.get_tasks()})
        names.sort()
        self.task_names = names
        self.combo_task_name.configure(values=self.task_names)
        if names:
            self.combo_task_name.set(names[0])
        else:
            self.combo_task_name.set("Sélectionnez une tâche")

    def refresh_task_list(self):
        self.refresh_task_names()
        for i in self.tree_tasks.get_children():
            self.tree_tasks.delete(i)
        for task in self.task_manager.get_tasks():
            assigned = task['assigned_to']
            if assigned:
                person, date, shift = assigned
                shift_str = f"{self.shifts[shift][0]}-{self.shifts[shift][1]}"
                date_str = date.isoformat()
            else:
                person, date_str, shift_str = "", "", ""
            self.tree_tasks.insert("", "end", iid=task["id"],
                                   values=(task["name"], task["duration"], person, date_str, shift_str))

    def add_task(self):
        name = self.combo_task_name.get()
        dur_text = self.entry_task_duration.get().strip()
        person = self.combo_person.get()
        date_text = self.entry_date.get().strip()
        shift_text = self.combo_shift.get()

        if name == "Sélectionnez une tâche":
            messagebox.showerror("Erreur", "Veuillez sélectionner une tâche")
            return

        try:
            duration = int(dur_text)
            if duration <= 0:
                raise ValueError()
        except:
            messagebox.showerror("Erreur", "Durée invalide")
            return

        try:
            date = datetime.datetime.strptime(date_text, "%Y-%m-%d").date()
        except:
            messagebox.showerror("Erreur", "Date invalide (format AAAA-MM-JJ)")
            return

        # Trouver shift index
        shift_index = None
        for idx, s in enumerate(self.shifts):
            st = f"{s[0]}-{s[1]}"
            if st == shift_text:
                shift_index = idx
                break
        if shift_index is None:
            messagebox.showerror("Erreur", "Shift invalide")
            return

        # Ajouter ou mettre à jour tâche (nouvelle tâche assignée)
        # Ici on cherche si tâche du même nom & durée existe non assignée, sinon création nouvelle
        existing_task_id = None
        for t in self.task_manager.get_tasks():
            if t["name"] == name and t["duration"] == duration and t["assigned_to"] is None:
                existing_task_id = t["id"]
                break

        if existing_task_id:
            self.task_manager.update_task(existing_task_id, assigned_to=(person, date, shift_index))
        else:
            self.task_manager.add_task(name, duration, (person, date, shift_index))

        self.entry_task_duration.delete(0, "end")
        self.refresh_task_list()
        self.refresh_planning_table()

    def delete_selected_task(self):
        selected = self.tree_tasks.selection()
        if not selected:
            messagebox.showinfo("Info", "Sélectionnez une tâche à supprimer")
            return
        for sel_id in selected:
            self.task_manager.delete_task(int(sel_id))
        self.refresh_task_list()
        self.refresh_planning_table()

    def save_rest_days(self):
        person = self.combo_rest_person.get()
        rest_str = self.entry_rest_days.get().strip()
        dates = []
        if rest_str:
            for d in rest_str.split(";"):
                d = d.strip()
                try:
                    date = datetime.datetime.strptime(d, "%Y-%m-%d").date()
                    dates.append(date)
                except:
                    messagebox.showerror("Erreur", f"Date invalide : {d}")
                    return
        self.rest_days[person] = dates
        self.refresh_planning_table()
        messagebox.showinfo("Succès", "Jours de repos enregistrés")

    def refresh_planning_table(self):
        # Vide et affiche planning des tâches assignées sur le mois actuel
        for widget in self.inner_frame.winfo_children():
            widget.destroy()

        self.lbl_month.configure(text=f"{self.current_month:02d}/{self.current_year}")

        days_in_month = (datetime.date(self.current_year, self.current_month % 12 + 1, 1) - datetime.timedelta(days=1)).day

        # Header : jours + shifts
        ctk.CTkLabel(self.inner_frame, text="Personne", width=100, anchor="w").grid(row=0, column=0, sticky="w", padx=2, pady=2)
        col = 1
        for day in range(1, days_in_month + 1):
            for shift_idx, shift in enumerate(self.shifts):
                lbl = ctk.CTkLabel(self.inner_frame, text=f"{day}\n{shift[0]}-{shift[1]}", width=60, fg_color="#444444", corner_radius=5)
                lbl.grid(row=0, column=col, padx=1, pady=1)
                col += 1

        # Lignes personnes
        for row_idx, person in enumerate(self.persons, start=1):
            ctk.CTkLabel(self.inner_frame, text=person, width=100, anchor="w", fg_color=self.person_colors.get(person, "#555555"), corner_radius=5).grid(row=row_idx, column=0, sticky="w", padx=2, pady=2)

            col_idx = 1
            for day in range(1, days_in_month + 1):
                date_obj = datetime.date(self.current_year, self.current_month, day)
                for shift_idx, shift in enumerate(self.shifts):
                    # Case vide par défaut
                    txt = ""
                    bg = "#222222"

                    # Si jour repos, fond rouge clair
                    if date_obj in self.rest_days.get(person, []):
                        bg = "#800000"

                    # Chercher tâche assignée ce jour / shift
                    for task in self.task_manager.get_tasks_for(person, date_obj, shift_idx):
                        txt = f"{task['name']} ({task['duration']}m)"
                        bg = "#004400"
                        break

                    lbl_cell = ctk.CTkLabel(self.inner_frame, text=txt, width=60, height=40, fg_color=bg, corner_radius=5)
                    lbl_cell.grid(row=row_idx, column=col_idx, padx=1, pady=1)
                    col_idx += 1

        self.inner_frame.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def prev_month(self):
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self.refresh_planning_table()

    def next_month(self):
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.refresh_planning_table()

    def export_excel(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not filename:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planning"

        ws["A1"] = "Personne"
        ws["B1"] = "Date"
        ws["C1"] = "Shift"
        ws["D1"] = "Tâche"
        ws["E1"] = "Durée"

        row = 2
        for task in self.task_manager.get_tasks():
            assigned = task.get("assigned_to")
            if assigned:
                person, date, shift_idx = assigned
                shift = self.shifts[shift_idx]
                ws.cell(row=row, column=1, value=person)
                ws.cell(row=row, column=2, value=date.isoformat())
                ws.cell(row=row, column=3, value=f"{shift[0]}-{shift[1]}")
                ws.cell(row=row, column=4, value=task["name"])
                ws.cell(row=row, column=5, value=task["duration"])
                row += 1

        try:
            wb.save(filename)
            messagebox.showinfo("Succès", "Planning exporté avec succès")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export: {e}")

if __name__ == "__main__":
    app = PlanningApp()
    app.mainloop()
